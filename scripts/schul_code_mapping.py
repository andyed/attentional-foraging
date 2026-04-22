"""Join iMotions numeric respondent IDs to external P/W study codes.

The Zenodo deposit does not ship a direct mapping. This module reconstructs it
by cross-referencing two signals that must be consistent for the same person:

  - iMotions side: fragment duration per (respondent × stimulus), from
    <SceneFragment> blocks in Data.xml.
  - xlsx side: sum of TimeSpent-G across all AOIs per (respondent × stimulus),
    from AOI metrics.xlsx.

The xlsx sum is time-gaze-inside-any-AOI, which must be ≤ the fragment duration
(total time that stimulus was on screen). So for the TRUE match, every per-stim
gap = xlsx_sum − fragment_duration is ≤ 0; for any mismatch, at least one Q will
blow through. This turns out to be uniquely identifying — verified 2026-04-22
on M1_Students against AOI metrics.xlsx.

Caveats:
  - Participants with very few fragments can be ambiguous; require ≥ 4 shared Qs.
  - Respondents excluded from xlsx (e.g. quality filtering) can't be matched.
  - Returns None for the respondent if no candidate passes the constraint.
"""
from __future__ import annotations

import os
from collections import defaultdict
from dataclasses import dataclass

from openpyxl import load_workbook

import schul_fragments


STUDY_LABELS = {  # Zenodo zip → "Study name" column value in xlsx
    "M1_Students.zip":   "M1",
    "M1_noStudents.zip": "M1",
    "M2_Students.zip":   "M2",
    "M2_noStudents.zip": "M2",
    "D1_Students.zip":   "D1",
    "D1_noStudents.zip": "D1",
    "D2_Students.zip":   "D2",
    "D2_noStudents.zip": "D2",
}


@dataclass
class MappingResult:
    respondent_id: int             # iMotions numeric ID
    external_code: str | None      # e.g. "P14", "W15", or None if no valid match
    total_abs_gap_ms: int          # sum of |xlsx_sum − fragment_duration| over shared Qs
    second_best_gap_ms: int        # gap between best and second-best candidates
    shared_q_count: int            # number of Q-labels used in the match
    candidate_scores: list[tuple[str, int]]  # sorted [(external_code, abs_gap_ms), ...]


def iMotions_durations(zip_path: str) -> dict[int, dict[str, int]]:
    """Return {respondent_id: {q_label: fragment_duration_ms}}."""
    xml = schul_fragments.read_data_xml(zip_path)
    stimuli = schul_fragments.parse_stimuli(xml)
    frags = schul_fragments.parse_fragments(xml)
    out: dict[int, dict[str, int]] = defaultdict(dict)
    for f in frags:
        if f.stim_id in stimuli:
            out[f.respondent_id][stimuli[f.stim_id].name] = f.duration_ms
    return dict(out)


def xlsx_dwell_totals(xlsx_path: str, study: str) -> dict[str, dict[str, int]]:
    """Return {external_code: {q_label: sum_of_TimeSpent-G_ms}} for one study.

    Header row 10 in 'AOI metrics' sheet; we skip past it with min_row=11.
    Column indices: 0=Study, 1=Stimulus, 2=Respondent, 8=TimeSpent-G (ms).
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["AOI metrics"]
    out: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in ws.iter_rows(min_row=11, values_only=True):
        if r[0] != study:
            continue
        name, q, ts = r[2], r[1], r[8]
        if isinstance(ts, int) and ts > 0 and name is not None and q is not None:
            out[name][q] += ts
    return {k: dict(v) for k, v in out.items()}


def score_candidate(imot: dict[str, int], xlsx: dict[str, int]) -> tuple[int, int, bool]:
    """Return (total_abs_gap_ms, shared_q_count, all_non_positive).

    all_non_positive = True iff xlsx_sum ≤ fragment_duration for every shared Q.
    Only candidates with all_non_positive=True should be accepted.
    """
    shared = set(imot) & set(xlsx)
    total = sum(abs(xlsx.get(q, 0) - imot[q]) for q in shared)
    positive_violation = any(xlsx.get(q, 0) > imot[q] for q in shared)
    return total, len(shared), not positive_violation


def build_mapping(
    zip_path: str,
    xlsx_path: str,
    min_shared_qs: int = 4,
) -> dict[int, MappingResult]:
    """Compute iMotions ID → external code mapping for one study zip."""
    zip_basename = os.path.basename(zip_path)
    study = STUDY_LABELS.get(zip_basename)
    if study is None:
        raise ValueError(f"unknown zip for study mapping: {zip_basename}")

    imot_by_rid = iMotions_durations(zip_path)
    xlsx_by_code = xlsx_dwell_totals(xlsx_path, study)

    results: dict[int, MappingResult] = {}
    for rid, imot_dict in imot_by_rid.items():
        scored = []
        for code, xlsx_dict in xlsx_by_code.items():
            total, n_shared, ok = score_candidate(imot_dict, xlsx_dict)
            if n_shared < min_shared_qs:
                continue
            scored.append((ok, total, n_shared, code))
        # Prefer constraint-satisfying candidates with lowest |gap|
        ok_scored = sorted([s for s in scored if s[0]], key=lambda s: s[1])
        if ok_scored:
            best = ok_scored[0]
            second = ok_scored[1][1] if len(ok_scored) > 1 else best[1]
            results[rid] = MappingResult(
                respondent_id=rid,
                external_code=best[3],
                total_abs_gap_ms=best[1],
                second_best_gap_ms=second,
                shared_q_count=best[2],
                candidate_scores=[(c[3], c[1]) for c in ok_scored[:5]],
            )
        else:
            results[rid] = MappingResult(
                respondent_id=rid, external_code=None,
                total_abs_gap_ms=0, second_best_gap_ms=0,
                shared_q_count=0, candidate_scores=[],
            )
    return results


if __name__ == "__main__":
    import argparse, json
    ap = argparse.ArgumentParser()
    ap.add_argument("zip", help="path to M1_Students.zip / etc")
    ap.add_argument("xlsx", help="path to AOI metrics.xlsx")
    ap.add_argument("--min-qs", type=int, default=4, help="minimum shared Q-labels required")
    args = ap.parse_args()

    mapping = build_mapping(args.zip, args.xlsx, min_shared_qs=args.min_qs)
    print(f"Mapping for {os.path.basename(args.zip)}:\n")
    unmapped = []
    for rid in sorted(mapping):
        r = mapping[rid]
        if r.external_code is None:
            unmapped.append(rid)
            continue
        gap_s = r.total_abs_gap_ms / 1000
        margin_s = (r.second_best_gap_ms - r.total_abs_gap_ms) / 1000
        tag = "✓" if margin_s >= 2 else ("~" if margin_s >= 0.5 else "??")
        print(f"  {rid} → {r.external_code:>5}  |gap| {gap_s:5.2f}s  margin {margin_s:5.2f}s  n={r.shared_q_count} {tag}")

    # Check injectivity
    codes = [r.external_code for r in mapping.values() if r.external_code]
    from collections import Counter
    dup = {c: n for c, n in Counter(codes).items() if n > 1}
    if dup:
        print(f"\n⚠ ambiguous codes claimed by multiple iMotions IDs: {dup}")
    if unmapped:
        print(f"\n⚠ unmapped iMotions IDs: {unmapped}")
    print(f"\n{len(codes)} of {len(mapping)} iMotions IDs uniquely matched ({len(set(codes))} distinct codes).")
